import smtplib
from email.message import EmailMessage
import pytest
from selenium import webdriver
import time
from fpdf import FPDF
import openpyxl

PDFtext = []
pdf = FPDF()
msg=EmailMessage()

# For Chrome Driver
driver=webdriver.Chrome(executable_path='C:\Python39\Lib\site-packages\selenium\webdriver\chrome\chromedriver')

#To maximize window
driver.maximize_window()
driver.implicitly_wait(10)
time.sleep(2)

# Working with Excel Sheet
path = "C:/Users/Gagan/PycharmProjects/Exercise_01/Test_script/PDF_01/python test.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for r in range(2,6):
    print("r is "+str(r))
    driver.get("https://beneficienttest.appiancloud.com/suite/")
    driver.find_element_by_xpath("//div[@class='login_box_inner']/form[2]/div[1]/div/input").send_keys(sheet.cell(row=r, column=1).value)
    driver.find_element_by_xpath("//div[@class='login_box_inner']/form[2]/div[2]/div/input").send_keys(sheet.cell(row=r, column=2).value)
    driver.find_element_by_xpath("//div[@class='login_box_inner']/form[2]/div[4]/div/div[2]/input").click()
    time.sleep(4)
    print(driver.title)

    if driver.title == "Funds - BIDS":
        print("Test Case is Passed")
        print(driver.title)

        sheet.cell(row=r,column=3).value="Test Case is Passed"
    else:
        print("Test Case is Failed")
        print(driver.title)
        sheet.cell(row=r,column=3).value="Test Case is Failed"
    workbook.save("C:/Users/Gagan/PycharmProjects/Exercise_01/Test_script/PDF_01/python test.xlsx")

driver.close()